# -*- coding: utf-8 -*-
"""Evaluate Admissibility.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1soCriVWlP8LnJxKUSwFLzA01VHrkNf_c
"""

!pip install python-docx openpyxl

# Import necessary libraries and modules
import hashlib      # Import hashlib for generating hashes
import json         # To work with JSON data structures
from openpyxl import load_workbook  # To handle Excel files
from docx import Document           # To handle Word files
from google.colab import files      # To download files from Colab
from IPython.display import display, clear_output  # To control IPython display
import ipywidgets as widgets        # For creating interactive widgets
import os                           # For OS related operations
from datetime import datetime       # To work with dates and times

# Functions

# Define a function to generate a hash from file content
def generate_hash(file_content):
    """Generate SHA256 hash of a file."""
    sha256 = hashlib.sha256()   # Create a new SHA256 hash object
    sha256.update(file_content)  # Update the hash object with the file content
    return sha256.hexdigest()   # Return the hex representation of the hash

# Define a function to extract metadata from a file path and file type
def get_metadata(file_path, file_type):
    """Retrieve enhanced metadata including additional features."""
    # Basic metadata extraction
    metadata = {
        'size_bytes': os.path.getsize(file_path),   # Get file size in bytes
        'last_access': datetime.fromtimestamp(os.path.getatime(file_path)).strftime('%Y-%m-%d %H:%M:%S')[:13]  # Last access timestamp
    }

        # If the file is a Word document, extract specific metadata
    if file_type == "word":
        doc = Document(file_path)
        num_images = len(doc.inline_shapes)  # Counting the number of images
        # Update the metadata dictionary with Word-specific metadata
        metadata.update({
            'author': doc.core_properties.author,
            'title': doc.core_properties.title,
            'created': str(doc.core_properties.created)[:13],
            'modified': str(doc.core_properties.modified)[:13],
            'word_count': sum([len(paragraph.text.split()) for paragraph in doc.paragraphs]),
            'num_images': num_images  # Additional Metadata
        })

    # Additional Metadata for Excel files
    elif file_type == "excel":
        wb = load_workbook(file_path, read_only=True)  # Load the Excel workbook
        num_sheets = len(wb.sheetnames)  # Count the number of sheets in the workbook
        sheet_names = wb.sheetnames  # Get the names of all sheets
        num_cells = sum([sum(1 for row in sheet.iter_rows()) for sheet in wb])  # Counting cells
        # Update the metadata dictionary with Excel-specific metadata
        metadata.update({
            'creator': wb.properties.creator,
            'title': wb.properties.title,
            'created': str(wb.properties.created)[:13],
            'modified': str(wb.properties.modified)[:13],
            'num_sheets': num_sheets,  # Additional Metadata
            'sheet_names': sheet_names,  # Additional Metadata
            'num_cells': num_cells  # Additional Metadata
        })

    return metadata   # Return the metadata dictionary

# Define a function to save forensic data as a JSON file
def save_forensic_data(data):
    """Save forensic data to a JSON file."""
    with open("forensic_data.json", 'w') as f: # Open a file in write mode
        json.dump(data, f) # Dump the forensic data into the file in JSON format
    files.download("forensic_data.json")  # Download the file

# Define a function to execute when a file is uploaded
def on_upload_generate(change):
    """Function to be called when a file is uploaded."""
    uploaded_file_details = list(change['new'].values())[0]
    uploaded_file_content = uploaded_file_details['content']
    uploaded_file_name = uploaded_file_details['metadata']['name']

    # Determine the file type
    file_type = 'word' if uploaded_file_name.endswith('.docx') else 'excel'

    # Generate hash value
    hash_value = generate_hash(uploaded_file_content)

    # Store the uploaded file temporarily
    file_extension = os.path.splitext(uploaded_file_name)[-1]
    temp_file_path = f"temp_file{file_extension}"
    with open(temp_file_path, "wb") as f:
        f.write(uploaded_file_content)

    try:
        # Retrieve metadata
        metadata = get_metadata(temp_file_path, file_type)
    except Exception as e:
       # If an error occurs during metadata extraction, display an error message and remove the temporary file
        clear_output(wait=True)
        display(file_uploader_generate)
        print(f"Error: {e}")
        os.remove(temp_file_path)
        return

    # Remove the temporary file after extracting metadata
    os.remove(temp_file_path)

    # Compile the forensic data and save it as a downloadable JSON file
    # Save and download forensic data
    forensic_data = {'hash': hash_value, 'metadata': metadata}
    save_forensic_data(forensic_data)
    clear_output(wait=True) # Clear the output
    display(file_uploader_generate) # Display the file uploader again
    print("Forensic data has been generated and downloaded.")  # Notify the user

# Create an interactive widget for file upload with a specific description and accepted file types
file_uploader_generate = widgets.FileUpload(description="Upload File", accept='.docx,.xlsx', multiple=False)
file_uploader_generate.observe(on_upload_generate, names='value') # Observe the value change of the widget and trigger the specified function

display(file_uploader_generate)   # Display the file upload widget

# # Define a function to load forensic data from a JSON content
def load_forensic_data(file_content):
    """Load forensic data from a JSON file."""
    return json.loads(file_content.decode("utf-8"))   # Decode and load JSON data from file content

# Define a function to execute when the "Compare" button is clicked
def on_upload_compare(change):
  # Extract forensic data details of the uploaded files
    forensic_file1_details = list(forensic_uploader1.value.values())[0]
    forensic_file2_details = list(forensic_uploader2.value.values())[0]

 # Load the forensic data of the two files
    forensic_data1 = load_forensic_data(forensic_file1_details['content'])
    forensic_data2 = load_forensic_data(forensic_file2_details['content'])

 # Check if the forensic data of the two files is identical
    admissible = forensic_data1 == forensic_data2

    clear_output(wait=True)   # Clear the output
     # Display the forensic data uploaders and the result text area
    display(forensic_uploader1, forensic_uploader2, result_text)

    # Update the result text area based on the comparison result
    if admissible:
        result_text.value = "The files are admissible."
    else:
        result_text.value = "The files are NOT admissible."

forensic_uploader1 = widgets.FileUpload(description="Upload Forensic JSON 1", accept='.json', multiple=False)
forensic_uploader2 = widgets.FileUpload(description="Upload Forensic JSON 2", accept='.json', multiple=False)

result_text = widgets.Textarea(value='', placeholder='', description='Result:', disabled=True)

compare_button = widgets.Button(description="Compare")
compare_button.on_click(on_upload_compare)

display(forensic_uploader1, forensic_uploader2, compare_button, result_text)